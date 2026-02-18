"""Explore team ranking data and decide: zones vs absolute rank."""
import openpyxl
import json
import sys
from collections import Counter, defaultdict

if sys.stdout.encoding != 'utf-8':
    sys.stdout.reconfigure(encoding='utf-8', errors='replace')

sys.path.insert(0, '.')
from config import PLAYER_DB_PATH

# Load team rankings from all sheets
wb = openpyxl.load_workbook("Teamrankingdata.xlsx")
team_ranks = {}  # (team, season_end_year) -> rank

for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    # Sheet name like "2009-2010" -> season ending 2010
    # But our draft years: 2009 draft = 2008-09 season
    # So "2008-2009" sheet corresponds to draft year 2009
    parts = sheet_name.split("-")
    season_end = int(parts[1])  # e.g. "2008-2009" -> 2009

    for r in range(2, ws.max_row + 1):
        rank = ws.cell(r, 1).value
        team = ws.cell(r, 2).value
        if rank and team:
            team_ranks[(team.strip(), season_end)] = int(rank)

print(f"Loaded {len(team_ranks)} team-season rankings")

# Load player DB
with open(PLAYER_DB_PATH) as f:
    db = json.load(f)

clean = [p for p in db if p.get("has_college_stats")
         and 2009 <= (p.get("draft_year") or 0) <= 2019
         and p.get("nba_ws") is not None]

# Match players to team ranks
# Draft year 2010 = 2009-2010 season = season_end 2010
matched = []
unmatched = []
for p in clean:
    college = p.get("college", "")
    draft_yr = p.get("draft_year", 0)
    rank = team_ranks.get((college, draft_yr))
    if rank is None:
        # Try common name variations
        for alt in [college.replace("St.", "State"), college.replace("State", "St."),
                    college + ".", college.rstrip(".")]:
            rank = team_ranks.get((alt, draft_yr))
            if rank:
                break
    if rank:
        matched.append({"name": p["name"], "tier": p["tier"], "rank": rank,
                        "college": college, "year": draft_yr, "level": p.get("level", "?")})
    else:
        unmatched.append({"name": p["name"], "college": college, "year": draft_yr})

print(f"Matched: {len(matched)}/{len(clean)} players")
if unmatched:
    print(f"Unmatched: {len(unmatched)}")
    for u in unmatched[:15]:
        print(f"  {u['name']:25s} {u['college']:20s} ({u['year']})")

# Key problem cases
print(f"\n=== KEY PLAYERS ===")
for name in ["Dominique Jones", "Gordon Hayward", "Jimmy Butler", "Stephen Curry",
             "Ja Morant", "Paul George", "Damian Lillard", "Jimmer Fredette",
             "Zion Williamson", "Trae Young", "Anthony Davis", "Kawhi Leonard"]:
    m = next((x for x in matched if x["name"] == name), None)
    if m:
        print(f"  {name:25s} {m['college']:20s} ({m['year']}) rank={m['rank']:>3}  level={m['level']}")

# ANALYSIS: Does team rank correlate with NBA tier?
print(f"\n\n=== TEAM RANK vs NBA TIER ===")
for t in range(1, 6):
    group = [m for m in matched if m["tier"] == t]
    if group:
        ranks = [m["rank"] for m in group]
        avg = sum(ranks) / len(ranks)
        ranks.sort()
        med = ranks[len(ranks)//2]
        top25 = sum(1 for r in ranks if r <= 25)
        top50 = sum(1 for r in ranks if r <= 50)
        top100 = sum(1 for r in ranks if r <= 100)
        print(f"  Tier {t}: avg rank={avg:>5.0f}  median={med:>3}  top25={top25/len(group)*100:>4.0f}%  top50={top50/len(group)*100:>4.0f}%  top100={top100/len(group)*100:>4.0f}%  (n={len(group)})")

# ANALYSIS: What does current level system look like in ranks?
print(f"\n=== CURRENT LEVEL SYSTEM vs ACTUAL RANK ===")
for level in ["High Major", "Mid Major", "Low Major"]:
    group = [m for m in matched if m["level"] == level]
    if group:
        ranks = [m["rank"] for m in group]
        avg = sum(ranks) / len(ranks)
        ranks.sort()
        lo = ranks[0]
        hi = ranks[-1]
        q25 = ranks[len(ranks)//4]
        q75 = ranks[3*len(ranks)//4]
        print(f"  {level:12s}: avg={avg:>5.0f}  range=[{lo}-{hi}]  IQR=[{q25}-{q75}]  n={len(group)}")

# ZONES ANALYSIS: What zone boundaries would make sense?
print(f"\n\n=== ZONE ANALYSIS: What if we used rank buckets? ===")
zones = [
    (1, 15, "Elite (1-15)"),
    (16, 40, "Strong (16-40)"),
    (41, 80, "Average (41-80)"),
    (81, 150, "Below Avg (81-150)"),
    (151, 400, "Weak (151+)"),
]
for lo, hi, label in zones:
    group = [m for m in matched if lo <= m["rank"] <= hi]
    if group:
        tiers = Counter(m["tier"] for m in group)
        n = len(group)
        stars = (tiers.get(1, 0) + tiers.get(2, 0)) / n * 100
        busts = (tiers.get(4, 0) + tiers.get(5, 0)) / n * 100
        print(f"  {label:20s}: n={n:>3}  star={stars:>4.0f}%  bust={busts:>4.0f}%  T1={tiers.get(1,0):>2} T2={tiers.get(2,0):>2} T3={tiers.get(3,0):>3} T4={tiers.get(4,0):>2} T5={tiers.get(5,0):>3}")

# Continuous: correlation coefficient
import math
ranks_list = [m["rank"] for m in matched]
tiers_list = [m["tier"] for m in matched]
n = len(matched)
mean_r = sum(ranks_list) / n
mean_t = sum(tiers_list) / n
cov = sum((r - mean_r) * (t - mean_t) for r, t in zip(ranks_list, tiers_list)) / n
std_r = (sum((r - mean_r)**2 for r in ranks_list) / n) ** 0.5
std_t = (sum((t - mean_t)**2 for t in tiers_list) / n) ** 0.5
corr = cov / (std_r * std_t) if std_r * std_t else 0
print(f"\n  Correlation (rank vs tier): r = {corr:.3f}")
print(f"  (positive = higher rank number = worse tier, which is expected)")

# What about using continuous rank in the model?
# The issue: rank 1 vs rank 5 is a small gap, rank 100 vs 200 is meaningless
# log(rank) might be better — compresses the tail
print(f"\n=== LOG(RANK) vs TIER ===")
import math
for t in range(1, 6):
    group = [m for m in matched if m["tier"] == t]
    if group:
        log_ranks = [math.log(m["rank"]) for m in group]
        avg = sum(log_ranks) / len(group)
        print(f"  Tier {t}: avg log(rank) = {avg:.2f}  (= rank ~{math.exp(avg):.0f})")
