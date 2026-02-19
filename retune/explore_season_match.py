"""Explore season matching between CBR, DraftPicks, and Barttorvik.
Key question: does CBR Season map cleanly to bar CSV year?
Also find gap players (last season != draft year - 1)."""
import sys, os, csv
import openpyxl
from collections import Counter

sys.stdout.reconfigure(encoding='utf-8', errors='replace')
DATA_DIR = 'NewCleanData'

# Load CBR
wb = openpyxl.load_workbook(os.path.join(DATA_DIR, 'SportsRefClean.xlsx'))
ws = wb.active
headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]

cbr_players = []
for r in range(2, ws.max_row + 1):
    row = {headers[c]: ws.cell(r, c+1).value for c in range(len(headers))}
    season = str(row.get('Season', ''))
    # "2010-11" -> bar year 2011
    if '-' in season:
        parts = season.split('-')
        bar_year = int(parts[0]) + 1
    else:
        bar_year = None
    cbr_players.append({
        'name': row.get('Player', ''),
        'college': row.get('Draft College', ''),
        'team': row.get('Team', ''),  # actual team name in CBR
        'season': season,
        'bar_year': bar_year,
        'class': row.get('Class', ''),
        'pos': row.get('Pos', ''),
    })

# Load DraftPicks
wb2 = openpyxl.load_workbook(os.path.join(DATA_DIR, 'AllCollegeDraftPicks.xlsx'))
ws2 = wb2.active
headers2 = [ws2.cell(1, c).value for c in range(1, ws2.max_column + 1)]

draft_picks = {}
for r in range(2, ws2.max_row + 1):
    name = ws2.cell(r, 1).value
    draft_year = ws2.cell(r, headers2.index('Draft Year') + 1).value
    pick = ws2.cell(r, headers2.index('Pick') + 1).value
    rd = ws2.cell(r, headers2.index('Round') + 1).value
    college = ws2.cell(r, headers2.index('LastCollegeTeam') + 1).value
    if name and draft_year:
        draft_picks[name] = {
            'draft_year': int(draft_year), 'pick': pick, 'round': rd,
            'college': college
        }

# Find gap players
print("=" * 80)
print("GAP PLAYERS (last college season != draft year - 1)")
print("=" * 80)
gaps = []
for p in cbr_players:
    dp = draft_picks.get(p['name'])
    if dp and p['bar_year']:
        expected_bar_year = dp['draft_year']  # draft 2020 -> expected last season 2019-20 = bar 2020
        if p['bar_year'] != expected_bar_year:
            gap = expected_bar_year - p['bar_year']
            gaps.append({
                **p, 'draft_year': dp['draft_year'], 'pick': dp['pick'],
                'gap': gap
            })

print(f"Total gap players: {len(gaps)}")
for g in sorted(gaps, key=lambda x: -abs(x['gap'])):
    print(f"  {g['name']:30s} season={g['season']}  bar_yr={g['bar_year']}  draft={g['draft_year']}  gap={g['gap']:+d}  pick={g['pick']}  class={g['class']}")

# Check bar CSV availability for all CBR players
print(f"\n{'=' * 80}")
print("BAR CSV AVAILABILITY")
print("=" * 80)
bar_years_needed = Counter(p['bar_year'] for p in cbr_players if p['bar_year'])
print(f"Bar years needed:")
for yr in sorted(bar_years_needed.keys()):
    csv_file = f"{yr}bar.csv"
    exists = os.path.exists(os.path.join(DATA_DIR, csv_file))
    print(f"  {yr}: {bar_years_needed[yr]} players  file={csv_file}  exists={exists}")

# Full match test: try to match ALL CBR players to Barttorvik
print(f"\n{'=' * 80}")
print("FULL MATCH TEST: CBR -> Barttorvik (by name)")
print("=" * 80)

# Load all bar CSVs into lookup
bar_lookup = {}  # (name, year) -> row data
BAR_HEADERS = [
    'player_name', 'team', 'conf', 'GP', 'Min_per', 'ORtg', 'usg', 'eFG',
    'TS_per', 'ORB_per', 'DRB_per', 'AST_per', 'TO_per', 'FTM', 'FTA',
    'FT_per', 'twoPM', 'twoPA', 'twoP_per', 'TPM', 'TPA', 'TP_per',
    'blk_per', 'stl_per', 'ftr', 'yr', 'ht', 'num', 'porpag', 'adjoe',
    'pfr', 'year', 'pid', 'type', 'Rec Rank', 'ast/tov', 'rimmade',
    'rimmade+rimmiss', 'midmade', 'midmade+midmiss',
    'rimmade/(rimmade+rimmiss)', 'midmade/(midmade+midmiss)',
    'dunksmade', 'dunksmiss+dunksmade', 'dunksmade/(dunksmade+dunksmiss)',
    'pick', 'drtg', 'adrtg', 'dporpag', 'stops', 'bpm', 'obpm', 'dbpm',
    'gbpm', 'mp', 'ogbpm', 'dgbpm', 'oreb', 'dreb', 'treb', 'ast', 'stl',
    'blk', 'pts', 'role', '3p/100'
]

for yr in range(2008, 2026):
    csv_file = os.path.join(DATA_DIR, f"{yr}bar.csv")
    if not os.path.exists(csv_file):
        continue
    with open(csv_file, encoding='utf-8') as f:
        reader = csv.reader(f)
        for row in reader:
            if len(row) >= 2:
                name = row[0].strip()
                bar_lookup[(name, yr)] = True

matched = 0
unmatched = []
no_bar_file = 0
for p in cbr_players:
    if not p['bar_year']:
        continue
    if not os.path.exists(os.path.join(DATA_DIR, f"{p['bar_year']}bar.csv")):
        no_bar_file += 1
        continue
    if (p['name'], p['bar_year']) in bar_lookup:
        matched += 1
    else:
        unmatched.append(p)

total = len([p for p in cbr_players if p['bar_year']])
print(f"Total CBR players: {total}")
print(f"No bar file for year: {no_bar_file}")
print(f"Matched by name+year: {matched}/{total - no_bar_file}")
print(f"Unmatched: {len(unmatched)}")
if unmatched:
    print(f"\nUnmatched players:")
    for u in sorted(unmatched, key=lambda x: x['bar_year']):
        # Try to find close name matches
        close = [k[0] for k in bar_lookup if k[1] == u['bar_year'] and
                 (u['name'].split()[-1] in k[0] or u['name'].split()[0] in k[0])]
        hint = f" -> maybe: {close[0]}" if close else ""
        print(f"  {u['name']:30s} ({u['bar_year']}, {u['college']}){hint}")

# Draft year distribution in DraftPicks
print(f"\n{'=' * 80}")
print("OUTCOME COVERAGE: How many players have BRef NBA data?")
print("=" * 80)
sys.path.insert(0, '.')
from config import PROCESSED_DIR
import json

bref_path = os.path.join(PROCESSED_DIR, 'bref_draft_stats.json')
if os.path.exists(bref_path):
    with open(bref_path) as f:
        bref = json.load(f)
    bref_names = set(bref.keys())
    print(f"BRef players: {len(bref_names)}")

    for yr_range, label in [((2007, 2019), "2007-2019 (mature)"),
                             ((2020, 2021), "2020-2021 (developing)"),
                             ((2022, 2023), "2022-2023 (young)"),
                             ((2024, 2025), "2024-2025 (TBD)")]:
        yr_players = [p for p in cbr_players if p['bar_year'] and
                      yr_range[0] <= draft_picks.get(p['name'], {}).get('draft_year', 0) <= yr_range[1]]
        has_bref = sum(1 for p in yr_players if p['name'] in bref_names)
        print(f"  {label}: {has_bref}/{len(yr_players)} have BRef data")
