"""Compare FinalTierCorrection.xlsx player list vs player_db.json."""
import openpyxl
import json
import os
import sys

if sys.stdout.encoding != 'utf-8':
    sys.stdout.reconfigure(encoding='utf-8', errors='replace')

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
from config import PLAYER_DB_PATH

# Load xlsx names
wb = openpyxl.load_workbook("audit/FinalTierCorrection.xlsx")
ws = wb.active
xlsx_players = {}
for r in range(2, ws.max_row + 1):
    name = ws.cell(r, 1).value
    if name:
        name = name.strip()
        corrected = ws.cell(r, 10).value  # Corrected Tier column
        current = ws.cell(r, 8).value     # Current Tier column
        tier = int(corrected) if corrected else (int(current) if current else None)
        xlsx_players[name] = {
            "tier": tier,
            "year": ws.cell(r, 2).value,
            "pick": ws.cell(r, 3).value,
            "college": ws.cell(r, 4).value,
        }
xlsx_names = set(xlsx_players.keys())

# Load DB
with open(PLAYER_DB_PATH) as f:
    db = json.load(f)

# DB players with college stats + 2009-2019 + outcomes
db_clean = [p for p in db if p.get("has_college_stats")
            and 2009 <= (p.get("draft_year") or 0) <= 2019
            and p.get("nba_ws") is not None]
db_names = {p["name"] for p in db_clean}

# All DB players 2009-2019 (with or without college stats)
db_all_range = [p for p in db if 2009 <= (p.get("draft_year") or 0) <= 2019
                and p.get("nba_ws") is not None]
db_all_names = {p["name"] for p in db_all_range}

print(f"XLSX players: {len(xlsx_names)}")
print(f"DB clean (2009-19 + college + WS): {len(db_clean)} ({len(db_names)} unique)")
print(f"DB all (2009-19 + WS, any stats): {len(db_all_range)} ({len(db_all_names)} unique)")

in_both = xlsx_names & db_names
in_xlsx_only = xlsx_names - db_names
in_db_only = db_names - xlsx_names

print(f"\nIn both: {len(in_both)}")
print(f"In XLSX only (not in DB with college stats): {len(in_xlsx_only)}")
print(f"In DB only (not in XLSX): {len(in_db_only)}")

# Check if XLSX-only are in DB at all (maybe without college stats)
in_xlsx_but_no_college = xlsx_names - db_all_names
in_xlsx_no_college_but_in_db = (xlsx_names - db_names) & db_all_names

print(f"\n  XLSX-only that ARE in DB but no college stats: {len(in_xlsx_no_college_but_in_db)}")
print(f"  XLSX-only that are NOT in DB at all: {len(in_xlsx_but_no_college)}")

if in_xlsx_only:
    print(f"\n--- XLSX ONLY ({len(in_xlsx_only)} players) ---")
    for n in sorted(in_xlsx_only):
        info = xlsx_players[n]
        in_db_flag = " (in DB, no college)" if n in in_xlsx_no_college_but_in_db else " (NOT in DB)"
        print(f"  {n:30s} {info['year']}  pick {info['pick']:>3}  {info['college'] or '?':20s}{in_db_flag}")

if in_db_only:
    print(f"\n--- DB ONLY ({len(in_db_only)} players, not in XLSX) ---")
    for n in sorted(in_db_only):
        p = next(x for x in db_clean if x["name"] == n)
        s = p.get("stats", {})
        gp = s.get("gp", 0) or 0
        mpg = s.get("mpg", 0) or 0
        print(f"  {n:30s} {p.get('draft_year')}  pick {p.get('draft_pick','?'):>3}  T{p['tier']}  GP={gp:.0f} MPG={mpg:.1f}  {p.get('college','?')}")

# Tier mismatches
print(f"\n--- TIER MISMATCHES (in both, different tiers) ---")
mismatches = 0
for n in sorted(in_both):
    xlsx_tier = xlsx_players[n]["tier"]
    p = next(x for x in db_clean if x["name"] == n)
    db_tier = p["tier"]
    if xlsx_tier and xlsx_tier != db_tier:
        mismatches += 1
        print(f"  {n:30s} XLSX=T{xlsx_tier}  DB=T{db_tier}")
print(f"Total mismatches: {mismatches}")
