"""Explore new data files and understand merge strategy."""
import sys, os, csv, json
import openpyxl

sys.stdout.reconfigure(encoding='utf-8', errors='replace')
sys.path.insert(0, '.')

DATA_DIR = 'NewCleanData'

# 1. SportsRefClean — the drafted player roster
print("=" * 80)
print("1. SportsRefClean.xlsx (CBR - drafted players, final season)")
print("=" * 80)
wb = openpyxl.load_workbook(os.path.join(DATA_DIR, 'SportsRefClean.xlsx'))
ws = wb.active
headers_cbr = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
print(f"Columns ({len(headers_cbr)}): {headers_cbr}")
print(f"Total rows: {ws.max_row - 1}")

# Sample rows
print(f"\nFirst 5 rows:")
for r in range(2, min(7, ws.max_row + 1)):
    row = {headers_cbr[c]: ws.cell(r, c+1).value for c in range(len(headers_cbr))}
    print(f"  {row.get('Player', ''):25s} {str(row.get('Season', '')):>8s} {str(row.get('Draft College', '')):>25s} Pick={row.get('Pick', '?')} Pos={row.get('Pos', '?')} Class={row.get('Class', '?')}")

# Year distribution
from collections import Counter
years = []
for r in range(2, ws.max_row + 1):
    season = ws.cell(r, headers_cbr.index('Season') + 1).value
    if season:
        # "2010-11" -> draft year 2011
        parts = str(season).split('-')
        if len(parts) == 2:
            yr = int(parts[0]) + 1
            years.append(yr)
yr_counts = Counter(years)
print(f"\nDraft year distribution:")
for yr in sorted(yr_counts.keys()):
    print(f"  {yr}: {yr_counts[yr]} players")
print(f"  TOTAL: {sum(yr_counts.values())}")

# 2. AllCollegeDraftPicks
print(f"\n{'=' * 80}")
print("2. AllCollegeDraftPicks.xlsx")
print("=" * 80)
wb2 = openpyxl.load_workbook(os.path.join(DATA_DIR, 'AllCollegeDraftPicks.xlsx'))
ws2 = wb2.active
headers_draft = [ws2.cell(1, c).value for c in range(1, ws2.max_column + 1)]
print(f"Columns: {headers_draft}")
print(f"Total rows: {ws2.max_row - 1}")

# Year distribution
draft_years = []
for r in range(2, ws2.max_row + 1):
    dy = ws2.cell(r, headers_draft.index('Draft Year') + 1).value
    if dy:
        draft_years.append(int(dy))
dy_counts = Counter(draft_years)
print(f"\nDraft year distribution:")
for yr in sorted(dy_counts.keys()):
    print(f"  {yr}: {dy_counts[yr]} players")
print(f"  TOTAL: {sum(dy_counts.values())}")

# 3. Barttorvik CSV sample
print(f"\n{'=' * 80}")
print("3. Barttorvik CSV sample (2024bar.csv)")
print("=" * 80)
with open(os.path.join(DATA_DIR, '2024bar.csv'), encoding='utf-8') as f:
    reader = csv.reader(f)
    rows = list(reader)
print(f"Total rows: {len(rows)}")
print(f"Columns per row: {len(rows[0]) if rows else 0}")

# Check if first row is header or data
print(f"\nFirst row (header check):")
for i, val in enumerate(rows[0][:10]):
    print(f"  col{i}: {val}")

# 4. barheaders.xlsx
print(f"\n{'=' * 80}")
print("4. barheaders.xlsx (column definitions)")
print("=" * 80)
wb3 = openpyxl.load_workbook(os.path.join(DATA_DIR, 'barheaders.xlsx'))
ws3 = wb3.active
print(f"Headers mapping:")
for r in range(1, ws3.max_row + 1):
    col_idx = ws3.cell(r, 1).value
    col_name = ws3.cell(r, 2).value
    if col_idx is not None or col_name is not None:
        print(f"  {r:3d}: {col_idx} -> {col_name}")

# 5. CompleteTeamrankingdata.xlsx
print(f"\n{'=' * 80}")
print("5. CompleteTeamrankingdata.xlsx")
print("=" * 80)
wb4 = openpyxl.load_workbook(os.path.join(DATA_DIR, 'CompleteTeamrankingdata.xlsx'))
print(f"Sheets: {wb4.sheetnames}")
print(f"Total sheets: {len(wb4.sheetnames)}")
# Compare to old team rankings
old_sheets = ['2008-2009', '2009-2010', '2010-2011', '2011-2012', '2012-2013',
              '2013-2014', '2014-2015', '2015-2016', '2016-2017', '2017-2018',
              '2018-2019', '2019-2020', '2020-2021']
new_sheets = set(wb4.sheetnames)
new_only = new_sheets - set(old_sheets)
print(f"New sheets (not in old data): {sorted(new_only)}")

# 6. Cross-reference: how many CBR players can we find in Barttorvik?
print(f"\n{'=' * 80}")
print("6. CROSS-REFERENCE: CBR players vs Barttorvik")
print("=" * 80)

# Load all bar data for one year to test matching
cbr_2024 = []
wb_cbr = openpyxl.load_workbook(os.path.join(DATA_DIR, 'SportsRefClean.xlsx'))
ws_cbr = wb_cbr.active
for r in range(2, ws_cbr.max_row + 1):
    season = str(ws_cbr.cell(r, headers_cbr.index('Season') + 1).value or '')
    if season.startswith('2023'):  # 2023-24 season = 2024 draft
        name = ws_cbr.cell(r, headers_cbr.index('Player') + 1).value
        college = ws_cbr.cell(r, headers_cbr.index('Draft College') + 1).value
        cbr_2024.append({'name': name, 'college': college})

# Load 2024 bar
bar_2024 = {}
with open(os.path.join(DATA_DIR, '2024bar.csv'), encoding='utf-8') as f:
    reader = csv.reader(f)
    for row in reader:
        if len(row) >= 2:
            bar_2024[row[0].strip()] = row[1].strip()

matched = 0
unmatched = []
for p in cbr_2024:
    if p['name'] in bar_2024:
        matched += 1
    else:
        unmatched.append(p)

print(f"2024 draft class: {len(cbr_2024)} CBR players")
print(f"2024 Barttorvik: {len(bar_2024)} players")
print(f"Name matches: {matched}/{len(cbr_2024)}")
if unmatched:
    print(f"Unmatched ({len(unmatched)}):")
    for u in unmatched[:20]:
        print(f"  {u['name']:30s} ({u['college']})")
