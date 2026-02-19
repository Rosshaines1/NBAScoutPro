"""Build team_ranks.json from CompleteTeamrankingdata.xlsx (Barttorvik rankings).

Output: { "Team Name|YYYY": rank, ... }
Where YYYY is the season-ending year (= draft year).

Sheet "2009-2010" -> season ending 2010 -> draft year 2010.
"""
import openpyxl
import json
import os
import sys

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
from config import PROCESSED_DIR, NEW_DATA_DIR

TEAM_RANKS_FILE = os.path.join(NEW_DATA_DIR, "CompleteTeamrankingdata.xlsx")

wb = openpyxl.load_workbook(TEAM_RANKS_FILE)

team_ranks = {}
for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    # "2008-2009" -> season ending 2009 -> draft year 2009
    parts = sheet_name.split("-")
    if len(parts) != 2:
        print(f"  Skipping sheet: {sheet_name}")
        continue
    season_end = int(parts[1])

    for r in range(2, ws.max_row + 1):
        rank = ws.cell(r, 1).value
        team = ws.cell(r, 2).value
        if rank and team:
            key = f"{team.strip()}|{season_end}"
            team_ranks[key] = int(rank)

    count = ws.max_row - 1
    print(f"  {sheet_name} -> draft {season_end}: {count} teams")

out_path = os.path.join(PROCESSED_DIR, "team_ranks.json")
os.makedirs(PROCESSED_DIR, exist_ok=True)
with open(out_path, "w") as f:
    json.dump(team_ranks, f, indent=2)

print(f"\nSaved {len(team_ranks)} team-season rankings to {out_path}")

# Verify key cases
for team, year in [("South Florida", 2010), ("Butler", 2010), ("Marquette", 2011),
                   ("Duke", 2019), ("Murray St.", 2019), ("San Diego St.", 2010)]:
    key = f"{team}|{year}"
    print(f"  {team} ({year}): rank {team_ranks.get(key, 'NOT FOUND')}")
