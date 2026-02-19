"""Build the player database from NewCleanData sources.

V5: Uses curated data from three primary sources:
    1. AllCollegeDraftPicks.xlsx — roster of all drafted players (pick, round, draft year)
    2. SportsRefClean.xlsx — final college season metadata (position, class, season)
    3. Barttorvik CSVs (2010-2025bar.csv) — ALL stats (advanced + counting), the "gospel"
    4. CompleteTeamrankingdata.xlsx -> team_ranks.json — team rankings for quadrant
    5. bref_draft_stats.json — NBA outcomes (Win Shares, VORP, etc.)

Key changes from V4:
    - Barttorvik is the sole source for all statistics (both advanced and counting)
    - CBR (SportsRefClean) provides metadata only: position, class, season mapping
    - Cut 2007-2009 drafts (no Barttorvik data before 2010)
    - TBD tier for recent drafts (2022+)
    - No more archive.zip dependency
"""
import sys
import os
import json
import csv
import re
import unicodedata

import numpy as np
import openpyxl

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
from config import (
    NEW_DATA_DIR, PROCESSED_DIR, PLAYER_DB_PATH, POSITIONAL_AVGS_PATH,
    TEAM_RANKS_PATH, QUADRANT_RANGES, BAR_HEADERS,
    TIER_LABELS, TBD_DRAFT_YEAR, HIGH_MAJORS, MID_MAJORS,
)
from pipeline.height_parser import parse_height

BREF_PATH = os.path.join(PROCESSED_DIR, "bref_draft_stats.json")

# Load team rankings for quadrant assignment
_team_ranks = {}
if os.path.exists(TEAM_RANKS_PATH):
    with open(TEAM_RANKS_PATH) as _f:
        _team_ranks = json.load(_f)


def get_quadrant(team_name, draft_year):
    """Get team quadrant from Barttorvik rankings."""
    if not team_name or not draft_year:
        return None, None
    key = f"{team_name}|{draft_year}"
    rank = _team_ranks.get(key)
    if rank is None:
        return None, None
    for quad, (lo, hi) in QUADRANT_RANGES.items():
        if lo <= rank <= hi:
            return quad, rank
    return "Q4", rank


# Players to exclude (bad data, too few games, overseas)
REMOVE_PLAYERS = {
    "Joel Embiid",       # Injured at Kansas, bad college data
    "James Wiseman",     # Played one college game at Memphis
    "Patrick Beverley",  # Played in Europe, not real US college data
}

# Win Shares tier thresholds
WS_TIERS = {
    1: 80,    # Superstar
    2: 40,    # All-Star
    3: 20,    # Solid Starter
    4: 5,     # Role Player
}

# Manual tier corrections from user review (Feb 2026, two rounds)
# Round 1: 99 corrections on 2009-2019. Round 2: 78 corrections on 2010-2021.
TIER_OVERRIDES = {
    # Round 1 (original)
    "Andrew Nicholson": 4, "Austin Rivers": 3, "Avery Bradley": 3,
    "Ben Simmons": 2, "Bol Bol": 4, "Bradley Beal": 1,
    "Brandon Ingram": 2, "Brandon Knight": 3, "Cam Reddish": 4,
    "Chinanu Onuaku": 4, "Coby White": 3, "Cody Martin": 3,
    "Collin Sexton": 3, "Danny Green": 3, "Darius Garland": 2,
    "Darren Collison": 3, "De'Aaron Fox": 2, "Deandre Ayton": 3,
    "Derrick Favors": 3, "Derrick White": 3, "Devin Booker": 1,
    "Devonte' Graham": 3, "Dillon Brooks": 3, "Dion Waiters": 3,
    "Donovan Mitchell": 1, "Dwight Powell": 3, "Ed Davis": 4,
    "Elfrid Payton": 3, "Eric Bledsoe": 3, "Eric Paschall": 4,
    "Frank Kaminsky": 3, "Gary Trent Jr.": 3, "Grant Williams": 3,
    "Greg Monroe": 3, "Greivis Vasquez": 3, "Iman Shumpert": 3,
    "Ja Morant": 2, "Jabari Parker": 3, "Jae Crowder": 3,
    "Jakob Poeltl": 3, "Jalen Brunson": 1, "Jalen McDaniels": 3,
    "Jared Sullinger": 3, "Jaren Jackson Jr.": 2, "Jarrett Allen": 3,
    "Jarrett Culver": 4, "Jaylen Brown": 1, "Jayson Tatum": 1,
    "Jeff Withey": 5, "Jonathan Isaac": 3, "Jordan Bell": 4,
    "Jordan Poole": 3, "Josh Hart": 3, "Justin Anderson": 4,
    "Kelly Olynyk": 3, "Kentavious Caldwell-Pope": 3,
    "Kevin Porter Jr.": 3, "Kevon Looney": 4, "Klay Thompson": 1,
    "Kyle Anderson": 3, "Kyle Kuzma": 3, "Lance Stephenson": 3,
    "Landry Shamet": 3, "Larry Nance Jr.": 4, "Lauri Markkanen": 2,
    "Lonzo Ball": 2, "Luke Kennard": 4,
    "Malik Monk": 3, "Markelle Fultz": 3, "Mason Plumlee": 3,
    "Meyers Leonard": 3, "Michael Carter-Williams": 3,
    "Michael Porter Jr.": 2, "Mo Bamba": 5, "Montrezl Harrell": 3,
    "Moritz Wagner": 3, "Nassir Little": 4,
    "Nickeil Alexander-Walker": 3, "Nikola Vucevic": 2,
    "Norris Cole": 4, "OG Anunoby": 2,
    "Rodney Hood": 3, "Rui Hachimura": 3,
    "Sam Young": 5, "Shai Gilgeous-Alexander": 1,
    "Skal Labissiere": 4, "Steven Adams": 3, "Taj Gibson": 3,
    "Talen Horton-Tucker": 3, "Terrence Jones": 3,
    "Tristan Thompson": 3, "Ty Lawson": 3, "Tyler Herro": 2,
    "Zach Collins": 3, "Zach LaVine": 2, "Zion Williamson": 2,
    # Round 2 (new corrections from TierReviewAndPostionalChange.csv)
    "Anthony Edwards": 1, "Brandon Boston Jr.": 4, "Cade Cunningham": 1,
    "Cameron Thomas": 3, "Carsen Edwards": 4, "Cassius Winston": 4,
    "Charles Bassey": 4, "Chris Duarte": 4, "Chuma Okeke": 4,
    "Cole Anthony": 3, "Dalano Banton": 4, "Damyean Dotson": 4,
    "Daniel Oturu": 4, "DeMarcus Cousins": 2, "Dennis Smith Jr.": 4,
    "Denzel Valentine": 4, "Desmond Bane": 2, "Devin Vassell": 2,
    "Dylan Windler": 4, "Edmond Sumner": 4, "Elijah Hughes": 4,
    "Evan Mobley": 2, "Evan Turner": 3, "Frank Jackson": 4,
    "Frank Mason": 4, "Grant Riller": 4, "Herb Jones": 4,
    "Ignas Brazdeikis": 4, "Isaiah Canaan": 4, "Isaiah Livers": 4,
    "Isaiah Stewart": 3, "Jaden Springer": 4, "Jake Layman": 4,
    "Jalen Johnson": 3, "Jalen Suggs": 3, "James Bouknight": 4,
    "Jared Butler": 4, "Jarred Vanderbilt": 3, "Jason Preston": 4,
    "Jimmer Fredette": 4, "John Wall": 2, "Jordan McRae": 4,
    "Jordan Nwora": 4, "Josh Christopher": 4, "Josh Jackson": 4,
    "Kai Jones": 4, "Kessler Edwards": 4, "Kira Lewis Jr.": 4,
    "Kyle Guy": 4, "Lonnie Walker IV": 4, "Malachi Flynn": 4,
    "Malachi Richardson": 4, "Miles McBride": 3, "Miye Oni": 4,
    "Nico Mannion": 4, "Nik Stauskas": 4, "P.J. Washington": 3,
    "Patrick McCaw": 4, "Precious Achiuwa": 3, "R.J. Barrett": 3,
    "RJ Barrett": 3, "Reggie Perry": 4, "Romeo Langford": 4,
    "Saben Lee": 4, "Santi Aldama": 3, "Scottie Barnes": 2,
    "Shane Larkin": 4, "Sharife Cooper": 4, "Skylar Mays": 4,
    "Tre Mann": 4, "Troy Brown": 4, "Ty Jerome": 3,
    "Tyler Ennis": 4, "Tyler Ulis": 4, "Tyrese Haliburton": 1,
    "Tyrese Maxey": 2, "Udoka Azubuike": 4, "Vernon Carey Jr.": 4,
    "Wade Baldwin IV": 4,
}

# Position corrections from user review
POS_OVERRIDES = {
    "Chandler Parsons": "W",
    "Dejounte Murray": "G",
    "Delon Wright": "G",
    "Draymond Green": "B",
    "Isaiah Stewart": "B",
    "Jalen Suggs": "G",
    "Jamal Murray": "G",
    "John Wall": "G",
    "Jordan Clarkson": "G",
    "Julius Randle": "B",
    "Malcolm Brogdon": "G",
    "Marcus Smart": "G",
    "Pascal Siakam": "B",
    "Precious Achiuwa": "B",
    "Ty Jerome": "G",
}

# Name aliases: CBR name -> bar CSV name (for unmatched players)
BAR_NAME_ALIASES = {
    # Formal/nickname differences
    "Bam Adebayo": "Edrice Adebayo",
    "Mohamed Bamba": "Mo Bamba",
    "Cameron Thomas": "Cam Thomas",
    "Cameron Johnson": "Cam Johnson",
    "Herb Jones": "Herbert Jones",
    "Nah'Shon Hyland": "Bones Hyland",
    "Herbert Jones": "Herbert Jones",  # CBR uses "Herb Jones"
    "Maurice Harkless": "Moe Harkless",
    "Joe Young": "Joseph Young",
    "Kay Felder": "Kahlil Felder",
    "Kezie Okpala": "KZ Okpala",
    "Luguentz Dort": "Lu Dort",
    "Marcus Sasser": "M.J. Sasser",
    "Kameron Jones": "Kam Jones",
    "Vince Edwards": "Vincent Edwards",
    "Dewan Hernandez": "Dewan Huell",
    "GG Jackson II": "Gregory Jackson",
    "Raymond Spalding": "Ray Spalding",
    "Trey Thompkins III": "Howard Thompkins III",
    # Suffix differences (CBR has "Jr." but bar uses ", Jr." or drops it)
    "Nickeil Alexander-Walker": "Nickeil Alexander-Walke",
    "PJ Washington": "P.J. Washington",
    "Dennis Smith Jr.": "Dennis Smith, Jr.",
    "Wendell Carter Jr.": "Wendell Carter",
    "Marvin Bagley III": "Marvin Bagley",
    "Jaren Jackson Jr.": "Jaren Jackson",
    "Gary Trent Jr.": "Gary Trent",
    "Larry Nance Jr.": "Larry Nance",
    "Kelly Oubre Jr.": "Kelly Oubre",
    "Otto Porter Jr.": "Otto Porter",
    "Tim Hardaway Jr.": "Tim Hardaway",
    "Trey Murphy III": "Trey Murphy",
    "Lonnie Walker IV": "Lonnie Walker",
    "Troy Brown Jr.": "Troy Brown",
    "Jabari Smith Jr.": "Jabari Smith",
    "Robert Williams III": "Robert Williams",
    "Derrick Jones Jr.": "Derrick Jones",
    "Kevin Porter Jr.": "Kevin Porter",
    "Kenyon Martin Jr.": "Kenyon Martin",
    "Michael Porter Jr.": "Michael Porter",
    # Apostrophe/accent differences
    "DeAndre Bembry": "DeAndre' Bembry",
    "Hamady N'Diaye": "Hamady Ndiaye",
}

# BRef name aliases: BRef draft name -> CBR name
BREF_NAME_ALIASES = {
    "Bam Adebayo": "Edrice Adebayo",
    "P.J. Washington": "PJ Washington",
    "Cam Thomas": "Cameron Thomas",
    "Herb Jones": "Herbert Jones",
    "Bones Hyland": "Nah'Shon Hyland",
    "Trey Murphy III": "Trey Murphy",
    "Maurice Harkless": "Moe Harkless",
    "Mo Bamba": "Mohamed Bamba",
    "Nic Claxton": "Nicolas Claxton",
    "Devyn Marble": "Roy Devyn Marble",
    "Svi Mykhailiuk": "Sviatoslav Mykhailiuk",
    "Wes Iwundu": "Wesley Iwundu",
    "Kay Felder": "Kahlil Felder",
    "Jeff Taylor": "Jeffery Taylor",
    "Joe Young": "Joseph Young",
}

CLASS_YEAR_MAP = {"Fr": 1, "FR": 1, "So": 2, "SO": 2, "Jr": 3, "JR": 3, "Sr": 4, "SR": 4}
BAR_CLASS_MAP = {"Fr": 1, "So": 2, "Jr": 3, "Sr": 4}


def normalize_name(name):
    """Normalize a name for fuzzy matching."""
    name = name.strip()
    name = unicodedata.normalize("NFD", name)
    name = "".join(c for c in name if unicodedata.category(c) != "Mn")
    name = name.replace(".", "")
    name = re.sub(r'\s+(Jr|Sr|III|II|IV)$', '', name, flags=re.IGNORECASE)
    return name.lower()


def safe_float(val, default=0.0):
    """Safely convert to float."""
    if val is None or val == "" or val == "NA":
        return default
    try:
        v = float(val)
        return v if not (isinstance(v, float) and np.isnan(v)) else default
    except (ValueError, TypeError):
        return default


def assign_position(height_inches):
    if height_inches < 76:
        return "G"
    elif height_inches > 81:
        return "B"
    return "W"


def get_conf_level(conf):
    if conf in HIGH_MAJORS:
        return "High Major"
    if conf in MID_MAJORS:
        return "Mid Major"
    return "Low Major"


def assign_tier_ws(nba_ws, draft_year):
    """Assign tier from Win Shares, or TBD for recent picks."""
    if draft_year and draft_year >= TBD_DRAFT_YEAR:
        return 6, TIER_LABELS[6]
    if nba_ws is None:
        return 5, TIER_LABELS[5]
    if nba_ws > WS_TIERS[1]:
        return 1, TIER_LABELS[1]
    if nba_ws > WS_TIERS[2]:
        return 2, TIER_LABELS[2]
    if nba_ws > WS_TIERS[3]:
        return 3, TIER_LABELS[3]
    if nba_ws > WS_TIERS[4]:
        return 4, TIER_LABELS[4]
    return 5, TIER_LABELS[5]


def load_draft_picks():
    """Load AllCollegeDraftPicks.xlsx -> dict of {name: {draft_year, pick, round, college}}."""
    path = os.path.join(NEW_DATA_DIR, "AllCollegeDraftPicks.xlsx")
    wb = openpyxl.load_workbook(path)
    ws = wb.active
    headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]

    picks = {}
    for r in range(2, ws.max_row + 1):
        row = {headers[c]: ws.cell(r, c + 1).value for c in range(len(headers))}
        name = str(row.get("Player", "")).strip()
        if not name:
            continue
        picks[name] = {
            "draft_year": int(row["Draft Year"]) if row.get("Draft Year") else None,
            "pick": int(row["Pick"]) if row.get("Pick") else 61,
            "round": int(row["Round"]) if row.get("Round") else 2,
            "college": str(row.get("LastCollegeTeam", "")).strip(),
        }
    return picks


def load_cbr():
    """Load SportsRefClean.xlsx -> dict of {name: {season, team, class, pos, counting_stats}}."""
    path = os.path.join(NEW_DATA_DIR, "SportsRefClean.xlsx")
    wb = openpyxl.load_workbook(path)
    ws = wb.active
    headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]

    players = {}
    for r in range(2, ws.max_row + 1):
        row = {headers[c]: ws.cell(r, c + 1).value for c in range(len(headers))}
        name = str(row.get("Player", "")).strip()
        if not name:
            continue

        season = str(row.get("Season", ""))
        # "2010-11" -> bar year 2011
        bar_year = None
        if "-" in season:
            parts = season.split("-")
            try:
                bar_year = int(parts[0]) + 1
            except ValueError:
                pass

        players[name] = {
            "season": season,
            "bar_year": bar_year,
            "team": str(row.get("Team", "")).strip(),
            "draft_college": str(row.get("Draft College", "")).strip(),
            "pos": str(row.get("Pos", "")).strip(),
            "class": str(row.get("Class", "")).strip(),
        }
    return players


def load_bar_csvs():
    """Load all Barttorvik CSVs into a lookup: (normalized_name, year) -> row dict."""
    bar_lookup = {}  # (norm_name, year) -> {col: val}
    raw_lookup = {}  # (exact_name, year) -> {col: val}

    for yr in range(2010, 2026):
        csv_path = os.path.join(NEW_DATA_DIR, f"{yr}bar.csv")
        if not os.path.exists(csv_path):
            continue
        with open(csv_path, encoding="utf-8", errors="replace") as f:
            reader = csv.reader(f)
            count = 0
            for row in reader:
                if len(row) < len(BAR_HEADERS):
                    # Pad short rows
                    row += [""] * (len(BAR_HEADERS) - len(row))
                d = {BAR_HEADERS[i]: row[i].strip() if i < len(row) else ""
                     for i in range(len(BAR_HEADERS))}
                name = d["player_name"].strip()
                if not name:
                    continue
                norm = normalize_name(name)
                # Keep latest entry per (name, year) — some CSVs have duplicates
                raw_lookup[(name, yr)] = d
                bar_lookup[(norm, yr)] = d
                count += 1
        print(f"  Loaded {yr}bar.csv: {count} players")

    return bar_lookup, raw_lookup


def load_bref():
    """Load Basketball Reference draft stats."""
    if not os.path.exists(BREF_PATH):
        print(f"  WARNING: {BREF_PATH} not found. Run pipeline/scrape_bref.py first.")
        return {}
    with open(BREF_PATH) as f:
        bref = json.load(f)
    # Build lookup: name -> best entry (handle duplicates by keeping highest WS)
    bref_map = {}
    for p in bref:
        name = p["name"]
        if name not in bref_map or (p.get("nba_ws") or 0) > (bref_map[name].get("nba_ws") or 0):
            bref_map[name] = p
    print(f"  BRef draft picks loaded: {len(bref_map)}")
    return bref_map


def match_bar(name, bar_year, bar_lookup, raw_lookup):
    """Try to find a player in bar data with multiple strategies."""
    if not bar_year:
        return None

    # 1. Direct alias match
    alias = BAR_NAME_ALIASES.get(name)
    if alias:
        key = (alias, bar_year)
        if key in raw_lookup:
            return raw_lookup[key]

    # 2. Exact name match
    key = (name, bar_year)
    if key in raw_lookup:
        return raw_lookup[key]

    # 3. Normalized name match
    norm = normalize_name(name)
    key = (norm, bar_year)
    if key in bar_lookup:
        return bar_lookup[key]

    # 4. Try without suffix (Jr., III, etc.)
    clean = re.sub(r'\s+(Jr\.?|Sr\.?|III|II|IV)$', '', name, flags=re.IGNORECASE).strip()
    if clean != name:
        key = (clean, bar_year)
        if key in raw_lookup:
            return raw_lookup[key]
        norm_clean = normalize_name(clean)
        key = (norm_clean, bar_year)
        if key in bar_lookup:
            return bar_lookup[key]

    return None


def match_bref(cbr_name, bref_map):
    """Match CBR player name to BRef data."""
    # Direct match
    if cbr_name in bref_map:
        return bref_map[cbr_name]

    # Check BRef aliases (BRef name -> CBR name)
    for bref_name, cbr_alias in BREF_NAME_ALIASES.items():
        if cbr_alias == cbr_name and bref_name in bref_map:
            return bref_map[bref_name]

    # Fuzzy/normalized match
    norm = normalize_name(cbr_name)
    for bref_name, data in bref_map.items():
        if normalize_name(bref_name) == norm:
            return data

    return None


def build_player_db():
    """Main pipeline: merge all data sources into player database."""
    print("=" * 60)
    print("Loading data sources...")
    print("=" * 60)

    draft_picks = load_draft_picks()
    print(f"  Draft picks: {len(draft_picks)}")

    cbr = load_cbr()
    print(f"  CBR players: {len(cbr)}")

    print("\nLoading Barttorvik CSVs...")
    bar_lookup, raw_lookup = load_bar_csvs()
    print(f"  Total bar entries: {len(bar_lookup)}")

    bref_map = load_bref()

    # Merge draft picks + CBR (should be ~1:1)
    print("\n" + "=" * 60)
    print("Merging data sources...")
    print("=" * 60)

    players = []
    stats = {
        "total": 0, "skipped_no_bar_year": 0, "skipped_pre_2010": 0,
        "skipped_removed": 0, "bar_matched": 0, "bar_unmatched": 0,
        "bref_matched": 0, "bref_unmatched": 0, "tbd_count": 0,
        "tier_overrides": 0,
    }
    unmatched_bar = []
    unmatched_cbr = []

    for name, dp in draft_picks.items():
        stats["total"] += 1

        if name in REMOVE_PLAYERS:
            stats["skipped_removed"] += 1
            continue

        draft_year = dp["draft_year"]
        draft_pick = dp["pick"]

        # Get CBR metadata
        cbr_data = cbr.get(name)
        if not cbr_data:
            # Try normalized match
            norm = normalize_name(name)
            for cbr_name, cbr_d in cbr.items():
                if normalize_name(cbr_name) == norm:
                    cbr_data = cbr_d
                    break

        if not cbr_data:
            unmatched_cbr.append(name)
            continue

        bar_year = cbr_data["bar_year"]

        # Skip players without bar year or pre-2010
        if not bar_year:
            stats["skipped_no_bar_year"] += 1
            continue
        if bar_year < 2010:
            stats["skipped_pre_2010"] += 1
            continue

        # Match to Barttorvik
        bar_data = match_bar(name, bar_year, bar_lookup, raw_lookup)
        if not bar_data:
            stats["bar_unmatched"] += 1
            unmatched_bar.append((name, bar_year, cbr_data["team"]))
            continue
        stats["bar_matched"] += 1

        # Match to BRef for NBA outcomes
        bref_data = match_bref(name, bref_map)
        nba_ws = None
        nba_vorp = None
        nba_bpm = None
        if bref_data:
            stats["bref_matched"] += 1
            nba_ws = bref_data.get("nba_ws")
            nba_vorp = bref_data.get("nba_vorp")
            nba_bpm = bref_data.get("nba_bpm")
        else:
            stats["bref_unmatched"] += 1

        # Parse bar stats
        gp = safe_float(bar_data.get("GP"), 20)
        gp_div = gp if gp > 0 else 30

        # Height from bar (preferred) or CBR
        height = parse_height(bar_data.get("ht"))
        pos = assign_position(height)

        # Class year from CBR (has clean Fr/So/Jr/Sr labels)
        cbr_class = cbr_data.get("class", "")
        class_year = CLASS_YEAR_MAP.get(cbr_class, 4)
        # Also try bar 'yr' column as backup
        if class_year == 4 and cbr_class not in CLASS_YEAR_MAP:
            bar_yr = str(bar_data.get("yr", "")).strip()
            class_year = BAR_CLASS_MAP.get(bar_yr, 4)

        # Conference/level (legacy) and quadrant
        conf = str(bar_data.get("conf", ""))
        level = get_conf_level(conf)

        # Team name for quadrant lookup — try bar team first, then CBR
        bar_team = str(bar_data.get("team", "")).strip()
        quadrant, team_rank = get_quadrant(bar_team, draft_year)
        if quadrant is None:
            # Try CBR team name
            quadrant, team_rank = get_quadrant(cbr_data["team"], draft_year)
        if quadrant is None:
            # Try draft picks college name
            quadrant, team_rank = get_quadrant(dp["college"], draft_year)

        # === EXTRACT ALL STATS FROM BARTTORVIK (gospel) ===

        # Per-game counting stats (already per-game in bar)
        ppg = safe_float(bar_data.get("pts"))
        rpg = safe_float(bar_data.get("treb"))
        apg = safe_float(bar_data.get("ast"))
        spg = safe_float(bar_data.get("stl"))
        bpg = safe_float(bar_data.get("blk"))
        mpg = safe_float(bar_data.get("mp"), 30.0)
        oreb_pg = safe_float(bar_data.get("oreb"))
        dreb_pg = safe_float(bar_data.get("dreb"))

        # Percentages — bar stores eFG/TS as 0-100, but FT/3P/2P as 0-1 decimals
        efg = safe_float(bar_data.get("eFG"), 45.0)
        tp_pct = safe_float(bar_data.get("TP_per")) * 100    # bar stores as 0.361 → 36.1
        ft_pct = safe_float(bar_data.get("FT_per")) * 100    # bar stores as 0.861 → 86.1
        ts_per = safe_float(bar_data.get("TS_per"))
        two_p_pct = safe_float(bar_data.get("twoP_per")) * 100  # bar stores as 0.520 → 52.0

        # Advanced rate stats
        bpm_val = safe_float(bar_data.get("bpm"))
        obpm = safe_float(bar_data.get("obpm"))
        dbpm = safe_float(bar_data.get("dbpm"))
        usg = safe_float(bar_data.get("usg"))
        stl_per = safe_float(bar_data.get("stl_per"))
        blk_per = safe_float(bar_data.get("blk_per"))
        orb_per = safe_float(bar_data.get("ORB_per"))
        drb_per = safe_float(bar_data.get("DRB_per"))
        ast_per = safe_float(bar_data.get("AST_per"))
        to_per = safe_float(bar_data.get("TO_per"))
        ftr_val = safe_float(bar_data.get("ftr"))

        # Efficiency ratings
        ortg = safe_float(bar_data.get("ORtg"))
        adjoe = safe_float(bar_data.get("adjoe"))
        adrtg = safe_float(bar_data.get("adrtg"))
        porpag_val = safe_float(bar_data.get("porpag"))
        dporpag_val = safe_float(bar_data.get("dporpag"))

        # Season totals → per-game
        fta_raw = safe_float(bar_data.get("FTA"))
        ftm_raw = safe_float(bar_data.get("FTM"))
        stops_raw = safe_float(bar_data.get("stops"))
        rimmade_raw = safe_float(bar_data.get("rimmade"))
        rim_att_raw = safe_float(bar_data.get("rimmade+rimmiss"))
        tpa_raw = safe_float(bar_data.get("TPA"))
        tpm_raw = safe_float(bar_data.get("TPM"))
        two_pa_raw = safe_float(bar_data.get("twoPA"))
        two_pm_raw = safe_float(bar_data.get("twoPM"))

        fta = fta_raw / gp_div if fta_raw else 0
        ftm = ftm_raw / gp_div if ftm_raw else 0
        stops = stops_raw / gp_div if stops_raw else 0
        rimmade = rimmade_raw / gp_div if rimmade_raw else 0
        rim_att = rim_att_raw / gp_div if rim_att_raw else 0
        tpa = tpa_raw / gp_div if tpa_raw else 0
        tpm = tpm_raw / gp_div if tpm_raw else 0
        two_pa = two_pa_raw / gp_div if two_pa_raw else 0
        two_pm = two_pm_raw / gp_div if two_pm_raw else 0

        # Derived stats
        ato_val = safe_float(bar_data.get("ast/tov"), 1.0)
        tpg_val = apg / ato_val if ato_val > 0 else apg

        # Assign tier
        tier, outcome = assign_tier_ws(nba_ws, draft_year)
        if draft_year and draft_year >= TBD_DRAFT_YEAR:
            stats["tbd_count"] += 1

        player = {
            "name": name,
            "college": bar_team or cbr_data["team"],
            "pos": pos,
            "h": height,
            "w": 200,          # placeholder
            "ws": height + 4,  # placeholder
            "age": class_year,
            "level": level,
            "quadrant": quadrant or "Q1",
            "team_rank": team_rank,
            "ath": 2,          # placeholder
            "draft_pick": draft_pick,
            "draft_year": draft_year,
            "nba_ws": round(nba_ws, 1) if nba_ws is not None else None,
            "nba_vorp": nba_vorp,
            "nba_bpm": nba_bpm,
            "has_college_stats": True,
            "stats": {
                "ppg": round(ppg, 1), "rpg": round(rpg, 1), "apg": round(apg, 1),
                "spg": round(spg, 1), "bpg": round(bpg, 1),
                "fg": round(efg, 1), "threeP": round(tp_pct, 1), "ft": round(ft_pct, 1),
                "tpg": round(tpg_val, 1), "mpg": round(mpg, 1), "bpm": round(bpm_val, 1),
                "fta": round(fta, 2), "ftm": round(ftm, 2),
                "obpm": round(obpm, 1), "dbpm": round(dbpm, 1),
                "stl_per": round(stl_per, 1), "usg": round(usg, 1),
                "stops": round(stops, 2), "rimmade": round(rimmade, 2),
                "rim_att": round(rim_att, 2),
                "ts_per": round(ts_per, 2), "adjoe": round(adjoe, 1),
                "adrtg": round(adrtg, 1), "gp": round(gp, 0),
                "tpa": round(tpa, 2), "tpm": round(tpm, 2),
                "two_pa": round(two_pa, 2), "two_pm": round(two_pm, 2),
                "two_p_pct": round(two_p_pct, 3),
                "oreb": round(oreb_pg, 2), "dreb": round(dreb_pg, 2),
                "ftr": round(ftr_val, 1),
                "orb_per": round(orb_per, 1), "drb_per": round(drb_per, 1),
                "ast_per": round(ast_per, 1), "to_per": round(to_per, 1),
                "blk_per": round(blk_per, 1),
                "ortg": round(ortg, 1),
                "porpag": round(porpag_val, 2),
                "dporpag": round(dporpag_val, 2),
            },
            "outcome": outcome,
            "tier": tier,
        }
        players.append(player)

    # --- Apply manual tier corrections ---
    tier_override_count = 0
    tier_override_index = {}
    for override_name, override_tier in TIER_OVERRIDES.items():
        tier_override_index[override_name] = override_tier
        stripped = unicodedata.normalize("NFKD", override_name).encode("ascii", "ignore").decode("ascii")
        if stripped != override_name:
            tier_override_index[stripped] = override_tier

    for p in players:
        # Don't override TBD players
        if p["tier"] == 6:
            continue
        name = p["name"]
        stripped_name = unicodedata.normalize("NFKD", name).encode("ascii", "ignore").decode("ascii")
        new_tier = tier_override_index.get(name) or tier_override_index.get(stripped_name)
        if new_tier is not None and new_tier != p["tier"]:
            p["tier"] = new_tier
            p["outcome"] = TIER_LABELS[new_tier]
            tier_override_count += 1

    stats["tier_overrides"] = tier_override_count

    # --- Apply manual position corrections ---
    pos_override_count = 0
    for p in players:
        new_pos = POS_OVERRIDES.get(p["name"])
        if new_pos and new_pos != p["pos"]:
            p["pos"] = new_pos
            pos_override_count += 1
    stats["pos_overrides"] = pos_override_count

    # Print summary
    print(f"\n  Results:")
    print(f"    Total draft picks:     {stats['total']}")
    print(f"    Removed (bad data):    {stats['skipped_removed']}")
    print(f"    No bar year:           {stats['skipped_no_bar_year']}")
    print(f"    Pre-2010 (cut):        {stats['skipped_pre_2010']}")
    print(f"    Bar matched:           {stats['bar_matched']}")
    print(f"    Bar unmatched:         {stats['bar_unmatched']}")
    print(f"    BRef matched:          {stats['bref_matched']}")
    print(f"    BRef unmatched:        {stats['bref_unmatched']}")
    print(f"    TBD (draft {TBD_DRAFT_YEAR}+):     {stats['tbd_count']}")
    print(f"    Tier overrides:        {tier_override_count}/{len(TIER_OVERRIDES)}")
    print(f"    Position overrides:    {pos_override_count}/{len(POS_OVERRIDES)}")
    print(f"    TOTAL PLAYERS:         {len(players)}")

    if unmatched_cbr:
        print(f"\n  CBR unmatched ({len(unmatched_cbr)}):")
        for n in unmatched_cbr[:10]:
            print(f"    {n}")

    if unmatched_bar:
        print(f"\n  Bar unmatched ({len(unmatched_bar)}):")
        for n, yr, team in unmatched_bar:
            print(f"    {n:35s} (bar_yr={yr}, team={team})")

    return players


def compute_positional_averages(players):
    """Compute real positional averages from players WITH college stats."""
    from collections import defaultdict
    stats_by_pos = defaultdict(lambda: defaultdict(list))
    stat_keys = ["ppg", "rpg", "apg", "spg", "bpg", "fg", "threeP", "ft", "tpg", "mpg"]

    for p in players:
        if not p.get("has_college_stats"):
            continue
        pos = p["pos"]
        for key in stat_keys:
            val = p["stats"].get(key, 0)
            if val is not None:
                stats_by_pos[pos][key].append(val)

    avgs = {}
    for pos in ["G", "W", "B"]:
        avgs[pos] = {}
        for key in stat_keys:
            vals = stats_by_pos[pos][key]
            avgs[pos][key] = round(np.mean(vals), 1) if vals else 0.0
        avg_apg = avgs[pos]["apg"]
        avg_tpg = avgs[pos]["tpg"]
        avgs[pos]["ato"] = round(avg_apg / avg_tpg, 2) if avg_tpg > 0 else 1.0

    return avgs


def validate(players):
    """Run sanity checks on the output."""
    print("\n--- VALIDATION ---")

    for check_name in ["Cade Cunningham", "Zion Williamson", "Ja Morant",
                        "Luka Garza", "Trae Young", "Paolo Banchero"]:
        found = [p for p in players if check_name in p["name"]]
        if found:
            p = found[0]
            ppg = p["stats"]["ppg"]
            bpm = p["stats"]["bpm"]
            ws = p.get("nba_ws", "?")
            print(f"  {check_name:25s} PPG={ppg:5.1f} BPM={bpm:5.1f} WS={ws} "
                  f"Tier={p['tier']} Q={p.get('quadrant')} draft={p.get('draft_year')}")
        else:
            print(f"  WARNING: {check_name} not found")

    from collections import Counter
    tier_counts = Counter(p["tier"] for p in players)
    print(f"\n  Tier distribution:")
    for t in sorted(tier_counts):
        label = TIER_LABELS.get(t, "?")
        print(f"    Tier {t} ({label:25s}): {tier_counts[t]}")

    total = len(players)
    with_stats = sum(1 for p in players if p.get("has_college_stats"))
    print(f"\n  Total: {total} ({with_stats} with college stats)")

    # Draft year distribution
    yr_counts = Counter(p["draft_year"] for p in players if p.get("draft_year"))
    print(f"\n  Draft year distribution:")
    for yr in sorted(yr_counts):
        print(f"    {yr}: {yr_counts[yr]} players")


def main():
    os.makedirs(PROCESSED_DIR, exist_ok=True)
    players = build_player_db()
    pos_avgs = compute_positional_averages(players)

    print(f"\nExporting {len(players)} players to {PLAYER_DB_PATH}")
    with open(PLAYER_DB_PATH, "w") as f:
        json.dump(players, f, indent=2)

    print(f"Exporting positional averages to {POSITIONAL_AVGS_PATH}")
    with open(POSITIONAL_AVGS_PATH, "w") as f:
        json.dump(pos_avgs, f, indent=2)

    validate(players)
    print("\nDone!")


if __name__ == "__main__":
    main()
